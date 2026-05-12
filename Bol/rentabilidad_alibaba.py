"""
rentabilidad_alibaba.py

Main analysis script: finds high-demand bol.com products, sources them from
AliExpress, and produces an Excel with profitability + investment filter.

Usage:
    python rentabilidad_alibaba.py

Config via environment variables or editing CONFIGURACION below.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
import time
import urllib.parse
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from demanda_bol import buscar_bestsellers, CATEGORIAS
from aliexpress_zoeken import buscar_en_aliexpress, estimar_envio_china_nl
from excel_colores import agregar_leyenda

_DIR = Path(__file__).parent

# ──────────────────────────────────────────────
#  CONFIGURACION — editá estos valores a gusto
# ──────────────────────────────────────────────
CONFIGURACION = {
    # Categorías a analizar (None = todas)
    'categorias': None,  # ej: ['speelgoed', 'sport-outdoor']

    # Páginas de bestsellers a scrapear por categoría
    'paginas_bol': 2,

    # Rango de inversión deseada (en €). Filtra por: precio_ali × moq
    'inversion_min': float(os.environ.get('INVERSION_MIN', 100)),
    'inversion_max': float(os.environ.get('INVERSION_MAX', 1000)),

    # Margen neto mínimo para que aparezca en el Excel (€)
    'margen_minimo': float(os.environ.get('MARGEN_MIN', 3.0)),

    # Comisión de bol.com (fracción). ~12% promedio
    'comision_bol': 0.12,

    # Derechos de importación EU estimados (fracción sobre precio + flete)
    # Varía por categoría: electrónica 0%, juguetes 4.7%, ropa 12%
    # Usamos 5% como estimación conservadora general
    'arancel_eu': 0.05,

    # BTW (VAT) Países Bajos 21%
    'btw': 0.21,

    # Peso promedio estimado si no se conoce (kg)
    'peso_default_kg': 0.5,

    # Delay entre requests (segundos)
    'delay_bol': 2.5,
    'delay_ali': 4.0,

    # Archivo de salida
    'archivo_salida': str(_DIR / 'rentabilidad_alibaba_bol.xlsx'),
}

# ──────────────────────────────────────────────
#  CÁLCULO DE COSTOS
# ──────────────────────────────────────────────

def calcular_precio_landed(precio_ali: float, flete_china_nl: float,
                            arancel: float, btw: float) -> dict:
    """
    Calculates the fully-landed unit cost after shipping, duties and VAT.

    EU import chain:
        1. valor_en_aduana = precio_ali + flete_china_nl
        2. aranceles       = valor_en_aduana × arancel
        3. base_btw        = valor_en_aduana + aranceles
        4. btw_importacion = base_btw × btw
        5. landed_total    = base_btw + btw_importacion
    """
    valor_aduana = precio_ali + flete_china_nl
    aranceles    = round(valor_aduana * arancel, 2)
    base_btw     = valor_aduana + aranceles
    btw_imp      = round(base_btw * btw, 2)
    landed       = round(base_btw + btw_imp, 2)
    return {
        'flete_china_nl':   flete_china_nl,
        'aranceles_eu':     aranceles,
        'btw_importacion':  btw_imp,
        'costo_landed':     landed,
    }


def calcular_margen(precio_venta_bol: float, costo_landed: float,
                    comision_bol: float, flete_nl: float = 0.0) -> dict:
    """
    Calculates net margin per unit sold on bol.com.

    bol.com charges commission on the full selling price (incl. BTW).
    BTW is collected from the buyer — the seller passes it to the tax authority.
    For simplicity we treat the bol.com listed price as the gross price (incl. BTW).
    """
    # bol.com commission on listed price
    comision       = round(precio_venta_bol * comision_bol, 2)
    # Net received after commission
    neto_recibido  = round(precio_venta_bol - comision - flete_nl, 2)
    margen_neto    = round(neto_recibido - costo_landed, 2)
    margen_pct     = round((margen_neto / precio_venta_bol) * 100, 1) if precio_venta_bol else 0
    return {
        'comision_bol':  comision,
        'flete_nl':      flete_nl,
        'margen_neto':   margen_neto,
        'margen_pct':    margen_pct,
    }


def clasificar(margen_neto: float, margen_pct: float) -> str:
    if margen_neto >= 15 or margen_pct >= 30:
        return 'GANADOR'
    if margen_neto >= 7 or margen_pct >= 15:
        return 'POTENCIAL'
    return 'MARGINAL'


# ──────────────────────────────────────────────
#  EXCEL OUTPUT
# ──────────────────────────────────────────────

FILL_VERDE    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_AMARILLO = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_ROJO     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_GRIS     = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FILL_HEADER   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_HEADER   = Font(color='FFFFFF', bold=True)
FILL_MAP      = {'GANADOR': FILL_VERDE, 'POTENCIAL': FILL_AMARILLO, 'MARGINAL': FILL_ROJO}


def _colorear_hoja(ws, df: pd.DataFrame):
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    headers = [c.value for c in ws[1]]
    for i, (_, row) in enumerate(df.iterrows(), start=2):
        estado = row.get('Estado', 'MARGINAL')
        fill = FILL_MAP.get(estado, FILL_ROJO)
        for cell in ws[i]:
            cell.fill = fill
        if estado == 'GANADOR':
            for cell in ws[i]:
                cell.font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
    ws.freeze_panes = 'A2'


def _hoja_inversion(wb, df: pd.DataFrame, inv_min: float, inv_max: float):
    """Adds a second sheet filtered by investment range."""
    nombre = f"Inversión €{int(inv_min)}-€{int(inv_max)}"
    if nombre in wb.sheetnames:
        del wb[nombre]
    ws = wb.create_sheet(nombre)

    df_filtrado = df[
        (df['Inversion_Minima_EUR'].notna()) &
        (df['Inversion_Minima_EUR'] >= inv_min) &
        (df['Inversion_Minima_EUR'] <= inv_max)
    ].copy()

    if df_filtrado.empty:
        ws['A1'] = f'Sin productos con inversión entre €{inv_min} y €{inv_max}'
        return

    # Write header
    cols = list(df_filtrado.columns)
    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    for i, (_, row) in enumerate(df_filtrado.iterrows(), start=2):
        estado = row.get('Estado', 'MARGINAL')
        fill = FILL_MAP.get(estado, FILL_ROJO)
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=i, column=j, value=row[col])
            cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
    ws.freeze_panes = 'A2'

    print(f"  📊 Hoja '{nombre}': {len(df_filtrado)} productos")


# ──────────────────────────────────────────────
#  PIPELINE PRINCIPAL
# ──────────────────────────────────────────────

def ejecutar():
    cfg = CONFIGURACION
    print("=" * 60)
    print("🛒 ANÁLISIS ALIBABA → BOL.COM")
    print(f"   Inversión objetivo: €{cfg['inversion_min']} – €{cfg['inversion_max']}")
    print("=" * 60)

    # 1. Scrapear bestsellers en bol.com
    print("\n📦 Paso 1: Buscando productos de alta demanda en bol.com...")
    productos_bol = buscar_bestsellers(
        categorias=cfg['categorias'],
        paginas=cfg['paginas_bol'],
        delay=cfg['delay_bol'],
    )
    print(f"   ✅ {len(productos_bol)} productos únicos encontrados en bol.com")

    if not productos_bol:
        print("❌ No se encontraron productos. Revisá la conexión o los selectores.")
        return

    # 2. Buscar precios en AliExpress y calcular rentabilidad
    print(f"\n🔍 Paso 2: Buscando precios en AliExpress ({len(productos_bol)} productos)...")
    resultados = []
    no_encontrados = 0

    for i, prod in enumerate(productos_bol):
        nombre_nl  = prod['Nombre_NL']
        precio_bol = prod['Precio_Bol']

        if (i + 1) % 10 == 0:
            print(f"   → {i + 1}/{len(productos_bol)} procesados...")

        ali = buscar_en_aliexpress(nombre_nl, delay=cfg['delay_ali'])

        if not ali['encontrado'] or not ali['precio_min_eur']:
            no_encontrados += 1
            continue

        precio_ali  = ali['precio_min_eur']
        peso_kg     = cfg['peso_default_kg']
        flete_china = estimar_envio_china_nl(peso_kg)

        landed = calcular_precio_landed(
            precio_ali, flete_china, cfg['arancel_eu'], cfg['btw']
        )
        margen = calcular_margen(
            precio_bol, landed['costo_landed'], cfg['comision_bol']
        )

        if margen['margen_neto'] < cfg['margen_minimo']:
            continue

        moq             = ali['moq']
        inversion_min   = round(landed['costo_landed'] * moq, 2)
        roi_pct         = round((margen['margen_neto'] / landed['costo_landed']) * 100, 1)

        resultados.append({
            'Estado':              clasificar(margen['margen_neto'], margen['margen_pct']),
            'Categoria':           prod['Categoria'],
            'Nombre_NL':           nombre_nl,
            'Num_Reviews_Bol':     prod['Num_Reviews'],
            'Precio_Venta_Bol':    precio_bol,
            'Precio_Ali_EUR':      precio_ali,
            'Flete_China_NL':      flete_china,
            'Aranceles_EU':        landed['aranceles_eu'],
            'BTW_Importacion':     landed['btw_importacion'],
            'Costo_Landed':        landed['costo_landed'],
            'Comision_Bol':        margen['comision_bol'],
            'Margen_Neto_EUR':     margen['margen_neto'],
            'Margen_Pct':          margen['margen_pct'],
            'ROI_Pct':             roi_pct,
            'MOQ':                 moq,
            'Inversion_Minima_EUR': inversion_min,
            'Titulo_Ali':          ali['titulo_ali'],
            'Link_Bol':            prod['Link_Bol'],
            'Link_Ali':            ali['url_producto'],
        })

    print(f"   ✅ {len(resultados)} productos con margen positivo")
    print(f"   ⚠️  {no_encontrados} sin precio en AliExpress")

    if not resultados:
        print("\n❌ Sin resultados con margen suficiente. Probá bajar MARGEN_MIN o cambiar categorías.")
        return

    # 3. Generar Excel
    print(f"\n📊 Paso 3: Generando Excel...")
    orden_estado = {'GANADOR': 0, 'POTENCIAL': 1, 'MARGINAL': 2}
    df = pd.DataFrame(resultados)
    df['_orden'] = df['Estado'].map(orden_estado)
    df = df.sort_values(['_orden', 'Margen_Neto_EUR'], ascending=[True, False])
    df = df.drop(columns='_orden').reset_index(drop=True)

    with pd.ExcelWriter(cfg['archivo_salida'], engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Todos los productos')
        _colorear_hoja(writer.sheets['Todos los productos'], df)
        agregar_leyenda(writer.book)

        # Hoja filtrada por inversión
        _hoja_inversion(writer.book, df, cfg['inversion_min'], cfg['inversion_max'])

    print(f"\n✅ Excel guardado en: {cfg['archivo_salida']}")
    print(f"   🟢 GANADORES:  {(df['Estado'] == 'GANADOR').sum()}")
    print(f"   🟡 POTENCIAL:  {(df['Estado'] == 'POTENCIAL').sum()}")
    print(f"   🔴 MARGINAL:   {(df['Estado'] == 'MARGINAL').sum()}")
    print(f"   💰 En rango inversión €{cfg['inversion_min']}-€{cfg['inversion_max']}: "
          f"{((df['Inversion_Minima_EUR'] >= cfg['inversion_min']) & (df['Inversion_Minima_EUR'] <= cfg['inversion_max'])).sum()}")


if __name__ == '__main__':
    ejecutar()
