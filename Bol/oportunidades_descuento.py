import sys
sys.stdout.reconfigure(encoding='utf-8')

import pandas as pd
import re
import urllib.parse
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from envio_utils import calcular_envio_es_nl
from bol_scraper import buscar_en_bol

_DIR = Path(__file__).parent
ARCHIVO_CSV    = str(_DIR / 'product_2399_es.csv')
ARCHIVO_SALIDA = str(_DIR / 'oportunidades_descuento.xlsx')

# ── Configuración ──────────────────────────────────────────────────────────────
DESCUENTO_MIN_PCT  = 30    # % mínimo de descuento PVD sobre PVP_BIGBUY
MIN_STOCK          = 5
MAX_PESO_KG        = 5.0
MIN_PVP            = 8.0   # evitar productos demasiado baratos para reventa

# Comisiones Bol.com
COMISION_BOL = 0.12
FEE_FIJO_BOL = 1.00

# Verificar precio real en Bol para cada producto (más lento pero más preciso).
# Solo incluye productos donde se encontró el producto con confianza ALTA o MEDIA.
# Poner en False para análisis rápido usando solo el PVP de BigBuy como referencia.
VERIFICAR_PRECIOS_BOL = True
DELAY_BOL = 2.0

# ── Marcas con riesgo activo de denuncia IP en LATAM ──────────────────────────
# Estas marcas tienen equipos legales que denuncian listings en ML activamente.
# No es ilegal revender, pero ML puede bajar tu listing si la marca denuncia.
MARCAS_RIESGO_IP = {
    # Procter & Gamble
    'gillette', 'oral-b', 'braun', 'pampers', 'always', 'tampax',
    'pantene', 'head & shoulders', 'head and shoulders', 'herbal essences',
    'ariel', 'bold', 'fairy', 'flash', 'febreze', 'vicks', 'old spice',
    'secret', 'safeguard', 'hugo boss', 'gucci',  # P&G tiene licencias de fragancia
    # Unilever
    'dove', 'axe', 'lynx', 'rexona', 'sure', 'degree',
    'lipton', 'knorr', 'hellmann', 'vaseline', 'tresemme',
    'clear shampoo', 'sunsilk', 'lux', 'signal',
    # L'Oréal grupo completo
    "l'oreal", 'loreal', 'maybelline', 'garnier', 'lancome',
    "kiehl's", 'kiehls', 'nyx cosmetics', 'urban decay', 'it cosmetics',
    'vichy', 'la roche-posay', 'cerave', 'skinceuticals',
    # Colgate-Palmolive
    'colgate', 'palmolive', 'ajax', 'speed stick', 'softsoap',
    # Johnson & Johnson / Kenvue
    "johnson's", 'johnsons', 'neutrogena', 'band-aid', 'listerine',
    'aveeno', 'rogaine', 'carefree', 'stayfree', 'o.b.',
    # Luxury / moda (enforcers agresivos)
    'louis vuitton', 'gucci', 'prada', 'chanel', 'hermes', 'hermès',
    'burberry', 'versace', 'armani', 'dior', 'yves saint laurent',
    'michael kors', 'coach', 'kate spade', 'tory burch',
    # Relojes de lujo
    'rolex', 'omega', 'tag heuer', 'breitling', 'iwc', 'patek',
    # Deportes (enforcement variable, Nike y Under Armour muy activos)
    'nike', 'under armour', 'new balance', 'reebok', 'asics',
    # Big tech / electrónica
    'apple', 'sony', 'dyson', 'bose', 'beats',
    # Alimentación masiva
    'nestle', 'coca-cola', 'pepsi', 'heinz', 'kraft', 'kellogg',
    'ferrero', 'nutella', 'haribo',
    # Farmacia / salud
    'bayer aspirin', 'voltaren', 'bepanthen',
}


def detectar_riesgo_ip(nombre, brand=''):
    """Detecta si el nombre o marca del producto coincide con una marca de riesgo."""
    texto = (str(nombre) + ' ' + str(brand)).lower()
    for marca in MARCAS_RIESGO_IP:
        if marca in texto:
            return 'ALTO', marca.title()
    return 'BAJO', ''


def calcular_pvp_bol(pvd, envio, ganancia=5.0):
    """Precio sugerido en Bol incluyendo comisión e IVA."""
    pvp_sin_iva = (pvd + envio + ganancia + FEE_FIJO_BOL) / (1 - COMISION_BOL)
    return round(pvp_sin_iva * 1.21, 2)


def calcular_ganancia_bol(pvd, envio, precio_venta_con_iva):
    """Ganancia neta vendiendo en Bol al precio dado."""
    pvp_sin_iva  = precio_venta_con_iva / 1.21
    comision     = pvp_sin_iva * COMISION_BOL + FEE_FIJO_BOL
    return round(pvp_sin_iva - comision - pvd - envio, 2)


def clasificar(ganancia, riesgo):
    if riesgo == 'ALTO':
        return 'RIESGO_IP'
    if ganancia >= 15: return 'GANADOR'
    if ganancia >= 8:  return 'POTENCIAL'
    return 'MARGINAL'


def ejecutar():
    print("=" * 65)
    print("  💰 Oportunidades por Descuento — BigBuy")
    if VERIFICAR_PRECIOS_BOL:
        print("  🔍 Modo: verificación de precios reales en Bol.com")
    else:
        print("  ⚡ Modo rápido: usando PVP de BigBuy como referencia")
    print("=" * 65)

    df = pd.read_csv(ARCHIVO_CSV, sep=';', encoding='utf-8', low_memory=False)
    df.columns = [c.strip().lower() for c in df.columns]

    for col in ['pvd', 'pvp_bigbuy', 'price', 'stock', 'weight', 'width', 'height', 'depth']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Filtros base
    df = df[
        (df['pvp_bigbuy'] > 0) &
        (df['pvd']        > 0) &
        (df['stock']      >= MIN_STOCK) &
        (df['weight']     > 0) &
        (df['weight']     <= MAX_PESO_KG) &
        (df['pvp_bigbuy'] >= MIN_PVP)
    ].copy()

    df['descuento_pct'] = ((df['pvp_bigbuy'] - df['pvd']) / df['pvp_bigbuy'] * 100).round(1)
    df['margen_vs_precio_lista'] = ((df['price'] - df['pvd']) / df['price'] * 100).round(1)

    df = df[df['descuento_pct'] >= DESCUENTO_MIN_PCT].copy()
    print(f"\n📦 Productos con descuento >= {DESCUENTO_MIN_PCT}%: {len(df)}")

    resultados = []
    sin_match_bol = 0

    for i, (_, item) in enumerate(df.iterrows()):
        pvd    = item['pvd']
        peso   = item['weight']
        nombre = str(item['name'])
        brand  = str(item.get('brand', ''))
        ean    = str(item.get('ean13', ''))

        envio  = calcular_envio_es_nl(peso, item['width'], item['height'], item['depth'])
        riesgo, marca_riesgo = detectar_riesgo_ip(nombre, brand)

        pvp_min = round(((pvd + envio + FEE_FIJO_BOL) / (1 - COMISION_BOL)) * 1.21, 2)

        # Precio de referencia para calcular ganancias
        precio_real_bol = None
        titulo_bol = ''
        confianza_bol = ''

        if VERIFICAR_PRECIOS_BOL:
            if (i + 1) % 20 == 0:
                print(f"   → {i + 1}/{len(df)} verificados en Bol...")
            resultado_bol = buscar_en_bol(
                ean=ean if ean and ean != 'nan' else None,
                nombre=nombre,
                delay=DELAY_BOL,
            )
            confianza_bol = resultado_bol['confianza']
            if confianza_bol in ('ALTA', 'MEDIA'):
                precio_real_bol = resultado_bol['precio']
                titulo_bol = resultado_bol['titulo_bol']
            elif confianza_bol == 'NO_ENCONTRADO':
                sin_match_bol += 1
                # Producto no está en Bol — descartarlo si estamos en modo verificación
                continue

        # Ganancias estimadas
        ganancia_bol     = calcular_ganancia_bol(pvd, envio, item['pvp_bigbuy'])
        ganancia_a_lista = calcular_ganancia_bol(pvd, envio, item['price']) if item['price'] > 0 else None
        ganancia_real    = calcular_ganancia_bol(pvd, envio, precio_real_bol) if precio_real_bol else None

        # Estado basado en precio real si está disponible, sino en PVP BigBuy
        ganancia_para_estado = ganancia_real if ganancia_real is not None else ganancia_bol
        estado = clasificar(ganancia_para_estado, riesgo)

        link_bol = resultado_bol['url'] if VERIFICAR_PRECIOS_BOL and confianza_bol in ('ALTA', 'MEDIA') else \
                   f"https://www.bol.com/nl/s/?searchtext={urllib.parse.quote(nombre[:60])}"

        fila = {
            'Estado':              estado,
            'Riesgo_IP':           riesgo,
            'Marca_Riesgo':        marca_riesgo,
            'Categoria':           str(item.get('category', '')),
            'Nombre':              nombre,
            'Titulo_Bol':          titulo_bol,
            'Confianza_Match':     confianza_bol,
            'Brand_ID':            brand,
            'EAN':                 ean,
            'Stock':               int(item['stock']),
            'Peso_kg':             peso,
            'Costo_PVD_EUR':       round(pvd, 2),
            'PVP_BigBuy_EUR':      round(item['pvp_bigbuy'], 2),
            'Precio_Lista_EUR':    round(item['price'], 2) if item['price'] > 0 else '',
            'Precio_Real_Bol_EUR': round(precio_real_bol, 2) if precio_real_bol else '',
            'Descuento_PVD_%':     item['descuento_pct'],
            'Margen_Lista_%':      item['margen_vs_precio_lista'] if item['price'] > 0 else '',
            'Envio_ES_NL_EUR':     envio,
            'PVP_Min_Venta_EUR':   pvp_min,
            'Gan_a_PVP_BigBuy':    ganancia_bol,
            'Gan_a_Precio_Lista':  round(ganancia_a_lista, 2) if ganancia_a_lista is not None else '',
            'Gan_a_Precio_Real':   round(ganancia_real, 2) if ganancia_real is not None else '',
            'Link_Bol':            link_bol,
            'Imagen1':             item['image1'] if pd.notna(item.get('image1')) else '',
            'Imagen2':             item['image2'] if pd.notna(item.get('image2')) else '',
        }
        resultados.append(fila)

    if VERIFICAR_PRECIOS_BOL:
        print(f"   ✅ Productos con match en Bol: {len(resultados)}")
        print(f"   ❌ Sin match en Bol (descartados): {sin_match_bol}")

    if not resultados:
        print("⚠️  Sin resultados.")
        return

    df_final = pd.DataFrame(resultados)

    orden = {'GANADOR': 0, 'POTENCIAL': 1, 'MARGINAL': 2, 'RIESGO_IP': 3}
    df_final['_ord'] = df_final['Estado'].map(orden).fillna(2)
    df_final = (df_final
                .sort_values(['_ord', 'Gan_a_PVP_BigBuy'], ascending=[True, False])
                .drop(columns='_ord')
                .reset_index(drop=True))

    with pd.ExcelWriter(ARCHIVO_SALIDA, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name='Descuentos')
        _colorear(writer.sheets['Descuentos'], df_final)
        _leyenda(writer.book)

    g  = (df_final['Estado'] == 'GANADOR').sum()
    p  = (df_final['Estado'] == 'POTENCIAL').sum()
    m  = (df_final['Estado'] == 'MARGINAL').sum()
    ri = (df_final['Estado'] == 'RIESGO_IP').sum()

    print(f"\n✅ Excel generado: {ARCHIVO_SALIDA}")
    print(f"   🟢 GANADORES:   {g}")
    print(f"   🟡 POTENCIAL:   {p}")
    print(f"   🔴 MARGINAL:    {m}")
    print(f"   ⚫ RIESGO IP:   {ri}")
    if VERIFICAR_PRECIOS_BOL:
        print(f"\n💡 'Gan_a_Precio_Real' = ganancia contra el precio real encontrado en Bol (más confiable).")
    print(f"   'Gan_a_Precio_Lista' = ganancia contra el precio de lista público (techo teórico).")


# ── Formato Excel ──────────────────────────────────────────────────────────────
FILLS = {
    'GANADOR':   PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'POTENCIAL': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'MARGINAL':  PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'RIESGO_IP': PatternFill(start_color='808080', end_color='808080', fill_type='solid'),
    'HEADER':    PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
}


def _colorear(ws, df):
    for cell in ws[1]:
        cell.fill = FILLS['HEADER']
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        fill = FILLS.get(row.get('Estado', 'MARGINAL'), FILLS['MARGINAL'])
        font = Font(color='FFFFFF') if row['Estado'] == 'RIESGO_IP' else Font(bold=(row['Estado'] == 'GANADOR'))
        for cell in ws[i]:
            cell.fill = fill
            cell.font = font

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)
    ws.freeze_panes = 'A2'


def _leyenda(wb):
    if 'Leyenda' in wb.sheetnames:
        del wb['Leyenda']
    ws = wb.create_sheet('Leyenda')
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 75

    filas = [
        ('HEADER',    'Estado',           'Cómo interpretar cada fila'),
        ('GANADOR',   'GANADOR',          'Ganancia neta > €15'),
        ('POTENCIAL', 'POTENCIAL',        'Ganancia €8–€15. Vale la pena evaluar.'),
        ('MARGINAL',  'MARGINAL',         'Ganancia < €8. Puede ser rentable a precio de lista.'),
        ('RIESGO_IP', 'RIESGO_IP',        'Marca con enforcement activo. No es ilegal revender, pero pueden bajar el listing.'),
        ('HEADER',    '',                 ''),
        ('HEADER',    'Columna',          'Descripción'),
        ('HEADER',    'Confianza_Match',  'ALTA = EAN verificado en Bol. MEDIA = match por nombre con alta similitud. BAJA = match dudoso.'),
        ('HEADER',    'Precio_Real_Bol',  'Precio real encontrado en Bol.com para este producto. Vacío si no se encontró.'),
        ('HEADER',    'Gan_a_Precio_Real','Ganancia real si vendés al precio actual de Bol. La columna más confiable.'),
        ('HEADER',    'Gan_a_PVP_BigBuy', 'Ganancia estimada usando el PVP de BigBuy como precio de venta (referencia).'),
        ('HEADER',    'Gan_a_Lista',      'Ganancia si vendés al precio de lista público (techo teórico).'),
        ('HEADER',    'Descuento_%',      'Descuento del PVD sobre el PVP de BigBuy. A mayor %, mejor margen.'),
        ('HEADER',    'PVP_Min',          'Precio mínimo en Bol para no perder dinero (breakeven).'),
    ]

    for r_idx, (fill_key, col_a, col_b) in enumerate(filas, 1):
        fill = FILLS.get(fill_key, FILLS['HEADER'])
        for c, val in [(1, col_a), (2, col_b)]:
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.fill = fill
            cell.font = Font(
                color='FFFFFF' if fill_key in ('HEADER', 'RIESGO_IP') else '000000',
                bold=(fill_key == 'HEADER')
            )
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[r_idx].height = 22


if __name__ == '__main__':
    ejecutar()
