import sys
sys.stdout.reconfigure(encoding='utf-8')

import requests
import pandas as pd
import json
import re
import time
import random
import urllib.parse
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from excel_colores import colorear_excel

_DIR = Path(__file__).parent
ARCHIVO_SALIDA = str(_DIR / 'tendencias_alibaba.xlsx')

# ── Configuración ──────────────────────────────────────────────────────────────
MAX_PRODUCTOS_BOL   = 12   # por categoría de Bol.com
MAX_ALIBABA         = 3    # proveedores Alibaba por producto
GANANCIA_OBJETIVO   = 5.00 # EUR ganancia neta objetivo por unidad (para PVP sugerido)

CATEGORIAS_BOL = [
    ('Hogar',        'https://www.bol.com/nl/nl/l/bestsellers/?cat_id=8294'),
    ('Juguetes',     'https://www.bol.com/nl/nl/l/bestsellers/?cat_id=8290'),
    ('Deportes',     'https://www.bol.com/nl/nl/l/bestsellers/?cat_id=9218'),
    ('Belleza',      'https://www.bol.com/nl/nl/l/bestsellers/?cat_id=8298'),
    ('Bebe',         'https://www.bol.com/nl/nl/l/bestsellers/?cat_id=8296'),
]

# Aranceles EU de importación China por categoría (aproximados)
ARANCEL = {
    'Hogar':       0.037,
    'Juguetes':    0.047,
    'Deportes':    0.037,
    'Belleza':     0.037,
    'Bebe':        0.047,
    'default':     0.035,
}

# Peso estimado en kg por categoría (cuando no hay dato real)
PESO_EST = {
    'Hogar':       0.70,
    'Juguetes':    0.35,
    'Deportes':    0.55,
    'Belleza':     0.25,
    'Bebe':        0.45,
    'default':     0.50,
}

HEADERS_BOL = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'nl-NL,nl;q=0.9,en;q=0.8',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}

HEADERS_ALIBABA = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Referer': 'https://www.google.com/',
}

# Productos de ejemplo como fallback si Bol.com bloquea el scraping
PRODUCTOS_EJEMPLO = [
    {'Categoria': 'Hogar',       'Nombre': 'Silicone Kitchen Spatula Set',      'Precio_Bol': 14.99, 'Reviews':  650},
    {'Categoria': 'Hogar',       'Nombre': 'Bamboo Cutting Board Large',        'Precio_Bol': 24.99, 'Reviews':  430},
    {'Categoria': 'Hogar',       'Nombre': 'Stainless Steel Water Bottle 1L',   'Precio_Bol': 22.99, 'Reviews':  760},
    {'Categoria': 'Juguetes',    'Nombre': 'Fidget Spinner Set 3 Pieces',       'Precio_Bol':  9.99, 'Reviews':  980},
    {'Categoria': 'Juguetes',    'Nombre': 'Magnetic Drawing Board Kids',       'Precio_Bol': 12.99, 'Reviews':  760},
    {'Categoria': 'Juguetes',    'Nombre': 'Building Blocks Classic 500pcs',   'Precio_Bol': 29.99, 'Reviews': 1500},
    {'Categoria': 'Deportes',    'Nombre': 'Resistance Bands Set 5 Levels',     'Precio_Bol': 17.99, 'Reviews': 1100},
    {'Categoria': 'Deportes',    'Nombre': 'Yoga Mat Non-Slip 6mm',             'Precio_Bol': 24.99, 'Reviews':  870},
    {'Categoria': 'Belleza',     'Nombre': 'Jade Roller Facial Massager',       'Precio_Bol': 14.99, 'Reviews':  520},
    {'Categoria': 'Belleza',     'Nombre': 'Electric Face Cleansing Brush',     'Precio_Bol': 22.99, 'Reviews':  340},
    {'Categoria': 'Bebe',        'Nombre': 'Baby Sound Machine Sleep Aid',      'Precio_Bol': 29.99, 'Reviews':  690},
    {'Categoria': 'Bebe',        'Nombre': 'Silicone Teething Toys BPA-free',   'Precio_Bol': 12.99, 'Reviews':  410},
]


# ── Tipo de cambio ─────────────────────────────────────────────────────────────
def get_tipo_cambio_usd():
    try:
        r = requests.get('https://open.er-api.com/v6/latest/EUR', timeout=10)
        usd = r.json().get('rates', {}).get('USD', 1.08)
        return float(usd)
    except Exception:
        return 1.08


# ── Costos de importación China → NL ──────────────────────────────────────────
def calcular_importacion(precio_fob_eur, peso_kg, categoria='default', metodo='aereo', moq=1):
    """
    Full import cost breakdown: China → Netherlands, por unidad.

    El flete se calcula para todo el pedido (peso_kg × moq) y se divide por moq,
    lo que da el costo de flete real por unidad — muy diferente a calcular 1 solo.

    BTW de importación mostrada por separado: es recuperable para importadores
    registrados en NL (se descuenta del BTW cobrado en ventas).
    """
    moq = max(moq, 1)
    arancel_pct = ARANCEL.get(categoria, ARANCEL['default'])

    peso_total = peso_kg * moq  # peso del pedido completo

    if metodo == 'aereo':
        flete_total = max(peso_total * 6.50, 15.00)   # air cargo China→NL
    elif metodo == 'maritimo':
        flete_total = max(peso_total * 2.50, 40.00)   # LCL sea freight China→Rotterdam
    else:                                               # express DHL/FedEx
        flete_total = max(peso_total * 16.00, 25.00)

    flete_unit = round(flete_total / moq, 2)           # flete por unidad

    cif_unit   = precio_fob_eur + flete_unit
    arancel    = round(cif_unit * arancel_pct, 2)
    btw_import = round((cif_unit + arancel) * 0.21, 2) # 21% sobre CIF+arancel (recuperable)
    costo_unit = round(precio_fob_eur + flete_unit + arancel, 2)  # sin BTW

    return {
        'Flete_Unit':  flete_unit,
        'Arancel':     arancel,
        'BTW_Import':  btw_import,   # referencial — recuperable para importadores NL
        'Costo_Unit':  costo_unit,
    }


# ── Rentabilidad en Bol.com ────────────────────────────────────────────────────
def calcular_rentabilidad(costo_unit, precio_bol, ganancia_obj=GANANCIA_OBJETIVO):
    """
    Dada la unidad de costo (post importación) y el precio actual en Bol,
    calcula: comisión, ganancia real y precio sugerido.
    """
    pvp_sin_iva  = precio_bol / 1.21
    comision_bol = round(pvp_sin_iva * 0.12 + 1.00, 2)
    ganancia_real = round(pvp_sin_iva * 0.88 - 1.00 - costo_unit, 2)

    # Precio sugerido para alcanzar ganancia objetivo
    pvp_sug = round(((costo_unit + ganancia_obj + 1.00) / 0.88) * 1.21, 2)

    return {
        'Comision_Bol':  comision_bol,
        'Ganancia_Real': ganancia_real,
        'PVP_Sugerido':  pvp_sug,
    }


# ── Clasificación ──────────────────────────────────────────────────────────────
def clasificar(ganancia):
    if ganancia >= 10: return 'GANADOR'
    if ganancia >= 5:  return 'POTENCIAL'
    return 'MARGINAL'


# ── Scraper Bol.com ────────────────────────────────────────────────────────────
def scrape_bol_categoria(categoria, url, max_prod=MAX_PRODUCTOS_BOL):
    productos = []
    try:
        time.sleep(random.uniform(2, 4))
        r = requests.get(url, headers=HEADERS_BOL, timeout=15)
        if r.status_code != 200:
            print(f"    ⚠️  HTTP {r.status_code} — {categoria}")
            return productos

        soup = BeautifulSoup(r.text, 'html.parser')

        # Bol.com usa distintos selectores según versión de página
        items = (
            soup.select('[data-test="product-item"]') or
            soup.select('[data-selenium="product-card"]') or
            soup.select('.product-item--row') or
            soup.select('[class*="product-item"]')
        )

        # Fallback: buscar JSON-LD en la página
        if not items:
            for script in soup.find_all('script', type='application/ld+json'):
                try:
                    data = json.loads(script.string or '')
                    if isinstance(data, dict) and data.get('@type') == 'ItemList':
                        for el in data.get('itemListElement', [])[:max_prod]:
                            item = el.get('item', {})
                            name  = item.get('name', '')
                            price = item.get('offers', {}).get('price', 0)
                            url_p = item.get('url', '')
                            if name and float(price or 0) > 0:
                                productos.append({
                                    'Categoria':  categoria,
                                    'Nombre':     name,
                                    'Precio_Bol': float(price),
                                    'Reviews':    0,
                                    'Link_Bol':   url_p,
                                })
                except Exception:
                    pass

        for item in items[:max_prod]:
            try:
                name_el = (
                    item.select_one('[data-test="product-title"]') or
                    item.select_one('.product-title') or
                    item.select_one('h3') or
                    item.select_one('h2')
                )
                name = name_el.get_text(strip=True) if name_el else None
                if not name:
                    continue

                price_el = (
                    item.select_one('[data-test="price"]') or
                    item.select_one('.promo-price') or
                    item.select_one('.price-block__price') or
                    item.select_one('[class*="price"]')
                )
                precio = _parse_precio(price_el.get_text(strip=True) if price_el else '0')

                rev_el = (
                    item.select_one('[data-test="rating-count"]') or
                    item.select_one('.reviews-score')
                )
                reviews = int(re.sub(r'\D', '', rev_el.get_text(strip=True)) or 0) if rev_el else 0

                link_el = item.select_one('a[href]')
                link = ''
                if link_el:
                    href = link_el.get('href', '')
                    link = ('https://www.bol.com' + href) if href.startswith('/') else href

                productos.append({
                    'Categoria':  categoria,
                    'Nombre':     name,
                    'Precio_Bol': precio,
                    'Reviews':    reviews,
                    'Link_Bol':   link,
                })
            except Exception:
                continue

        print(f"    ✓ {categoria}: {len(productos)} productos")
    except Exception as e:
        print(f"    ❌ {categoria}: {e}")

    return productos


def _parse_precio(texto):
    texto = texto.replace(',', '.').replace('€', '').replace('\xa0', '').strip()
    m = re.search(r'(\d+\.?\d*)', texto)
    return float(m.group(1)) if m else 0.0


# ── Scraper Alibaba ────────────────────────────────────────────────────────────
def buscar_alibaba(termino, max_res=MAX_ALIBABA):
    """
    Busca proveedores en Alibaba con filtros Gold Supplier + Trade Assurance.
    Intenta scraping HTML; si es bloqueado, devuelve URL de búsqueda manual.
    """
    termino_clean = re.sub(r'[^\w\s]', '', termino)[:60]
    url_search = (
        'https://www.alibaba.com/trade/search'
        f'?SearchText={urllib.parse.quote(termino_clean)}'
        '&tab=all&isGold=y&ta=y&sortType=total_tranamt'
    )

    resultados = []
    try:
        time.sleep(random.uniform(3, 6))
        session = requests.Session()
        session.headers.update(HEADERS_ALIBABA)
        r = session.get(url_search, timeout=15)

        if r.status_code != 200:
            return [_fallback_alibaba(termino_clean, url_search)]

        soup = BeautifulSoup(r.text, 'html.parser')

        # Buscar JSON embebido en la página (window.__PAGE_DATA__ o similar)
        for script in soup.find_all('script'):
            content = script.string or ''
            for pattern in [r'window\.__PAGE_DATA__\s*=\s*(\{.+?\});\s*\n',
                            r'window\.__GLOBAL_DATA__\s*=\s*(\{.+?\});\s*\n']:
                m = re.search(pattern, content, re.DOTALL)
                if m:
                    try:
                        data = json.loads(m.group(1))
                        offers = (
                            data.get('data', {}).get('offerList', []) or
                            data.get('offerList', [])
                        )
                        for offer in offers[:max_res]:
                            resultados.append(_parse_offer_json(offer, url_search))
                        if resultados:
                            return resultados
                    except Exception:
                        pass

        # Fallback: selectores HTML de Alibaba
        items = (
            soup.select('.J-offer-wrapper') or
            soup.select('[class*="offer-card"]') or
            soup.select('[class*="product-card"]') or
            soup.select('[class*="item-info"]')
        )

        for item in items[:max_res]:
            try:
                name_el  = item.select_one('[class*="title"]') or item.select_one('h3')
                price_el = item.select_one('[class*="price"]')
                moq_el   = (item.select_one('[class*="moq"]') or
                             item.select_one('[class*="min-order"]') or
                             item.select_one('[class*="min_order"]'))
                link_el  = item.select_one('a[href]')

                name  = name_el.get_text(strip=True)[:80] if name_el else termino_clean
                price = _parse_usd(price_el.get_text(strip=True) if price_el else '0')
                moq   = _parse_moq(moq_el.get_text(strip=True) if moq_el else '1')
                href  = link_el.get('href', url_search) if link_el else url_search
                link  = ('https:' + href) if href.startswith('//') else href

                is_gold = bool(item.select('[class*="gold"]') or item.select('[class*="Gold"]'))
                is_ta   = bool(item.select('[class*="trade-assurance"]') or item.select('[class*="ta"]'))

                resultados.append({
                    'Proveedor':       name,
                    'Precio_USD':      price,
                    'MOQ':             moq,
                    'Gold_Supplier':   '✓' if is_gold else '?',
                    'Trade_Assurance': '✓' if is_ta   else '?',
                    'Link_Alibaba':    link,
                    'Nota':            '',
                })
            except Exception:
                continue

        return resultados if resultados else [_fallback_alibaba(termino_clean, url_search)]

    except Exception as e:
        print(f"      ⚠️  Alibaba ({termino_clean[:30]}): {e}")
        return [_fallback_alibaba(termino_clean, url_search)]


def _parse_offer_json(offer, fallback_url):
    precio_info = offer.get('priceInfo', offer.get('price', {}))
    if isinstance(precio_info, dict):
        precio_str = precio_info.get('price', precio_info.get('minPrice', '0'))
    else:
        precio_str = str(precio_info)

    return {
        'Proveedor':       str(offer.get('companyName', offer.get('title', '')))[:80],
        'Precio_USD':      _parse_usd(str(precio_str)),
        'MOQ':             _parse_moq(str(offer.get('moq', '1'))),
        'Gold_Supplier':   '✓' if offer.get('isGold') else '?',
        'Trade_Assurance': '✓' if offer.get('tradeAssurance') else '?',
        'Link_Alibaba':    offer.get('detailUrl', fallback_url),
        'Nota':            '',
    }


def _fallback_alibaba(termino, url):
    return {
        'Proveedor':       'Ver manualmente en Alibaba',
        'Precio_USD':      0.0,
        'MOQ':             0,
        'Gold_Supplier':   'N/D',
        'Trade_Assurance': 'N/D',
        'Link_Alibaba':    url,
        'Nota':            'Scraping bloqueado — abrir link para ver precios',
    }


def _parse_usd(texto):
    """Extrae precio mínimo de rangos como '$2.50 - $8.00'."""
    texto = re.sub(r'[^\d.,]', '', texto.split('-')[0].split('~')[0])
    texto = texto.replace(',', '.')
    m = re.search(r'(\d+\.?\d*)', texto)
    return float(m.group(1)) if m else 0.0


def _parse_moq(texto):
    m = re.search(r'(\d+)', str(texto))
    return int(m.group(1)) if m else 1


# ── Main ───────────────────────────────────────────────────────────────────────
def ejecutar():
    print("=" * 65)
    print("  🔍 Tendencias Bol.com × Alibaba — Análisis de Importación")
    print("=" * 65)

    tipo_cambio = get_tipo_cambio_usd()
    print(f"\n💱 Tipo de cambio: 1 EUR = {tipo_cambio:.4f} USD\n")

    # ── 1. Scraping Bol.com ────────────────────────────────────────────────────
    print("📦 Obteniendo bestsellers de Bol.com...")
    productos_bol = []
    for cat, url in CATEGORIAS_BOL:
        productos_bol.extend(scrape_bol_categoria(cat, url))

    if not productos_bol:
        print("\n⚠️  Bol.com bloqueó el scraping. Usando productos de referencia.")
        productos_bol = PRODUCTOS_EJEMPLO
    else:
        # Filtrar productos sin precio y ordenar por popularidad
        productos_bol = [p for p in productos_bol if p['Precio_Bol'] > 0]
        productos_bol.sort(key=lambda x: x['Reviews'], reverse=True)

    print(f"\n✅ Productos a analizar: {len(productos_bol)}\n")

    # ── 2. Búsqueda Alibaba + análisis de importación ─────────────────────────
    print("🌏 Buscando proveedores en Alibaba...\n")
    resultados = []
    total = len(productos_bol)

    for i, prod in enumerate(productos_bol, 1):
        nombre     = prod['Nombre']
        precio_bol = prod['Precio_Bol']
        categoria  = prod['Categoria']
        peso_kg    = PESO_EST.get(categoria, PESO_EST['default'])

        print(f"  [{i:02d}/{total}] {nombre[:55]}")
        proveedores = buscar_alibaba(nombre)

        for prov in proveedores:
            precio_fob_usd = prov['Precio_USD']
            precio_fob_eur = round(precio_fob_usd / tipo_cambio, 2) if precio_fob_usd > 0 else 0.0

            moq = prov['MOQ'] if prov['MOQ'] > 0 else 50  # default conservador si no hay dato

            if precio_fob_eur > 0:
                imp_a = calcular_importacion(precio_fob_eur, peso_kg, categoria, 'aereo',    moq)
                imp_m = calcular_importacion(precio_fob_eur, peso_kg, categoria, 'maritimo', moq)
                imp_e = calcular_importacion(precio_fob_eur, peso_kg, categoria, 'express',  moq)

                rent_a = calcular_rentabilidad(imp_a['Costo_Unit'], precio_bol)
                rent_m = calcular_rentabilidad(imp_m['Costo_Unit'], precio_bol)

                estado = clasificar(rent_a['Ganancia_Real'])
            else:
                imp_a = imp_m = imp_e = {'Flete_Unit': 0, 'Arancel': 0, 'BTW_Import': 0, 'Costo_Unit': 0}
                rent_a = rent_m = {'Comision_Bol': 0, 'Ganancia_Real': 0, 'PVP_Sugerido': 0}
                estado = 'N/D'

            resultados.append({
                'Estado':              estado,
                'Categoria':           categoria,
                'Producto_Bol':        nombre,
                'Precio_Bol_EUR':      precio_bol,
                'Reviews_Bol':         prod.get('Reviews', 0),
                'Proveedor_Alibaba':   prov['Proveedor'],
                'MOQ':                 moq,
                'Gold_Supplier':       prov['Gold_Supplier'],
                'Trade_Assurance':     prov['Trade_Assurance'],
                'Precio_FOB_USD':      round(precio_fob_usd, 2),
                'Precio_FOB_EUR':      precio_fob_eur,
                'Peso_Est_kg':         peso_kg,
                'Flete_x_Unit_Aereo':  imp_a['Flete_Unit'],
                'Flete_x_Unit_Mar':    imp_m['Flete_Unit'],
                'Flete_x_Unit_Expr':   imp_e['Flete_Unit'],
                'Arancel_EUR':         imp_a['Arancel'],
                'BTW_Import_EUR':      imp_a['BTW_Import'],
                'Costo_Unit_Aereo':    imp_a['Costo_Unit'],
                'Costo_Unit_Maritimo': imp_m['Costo_Unit'],
                'Costo_Unit_Express':  imp_e['Costo_Unit'],
                'Comision_Bol_EUR':    rent_a['Comision_Bol'],
                'Ganancia_Aereo_EUR':  rent_a['Ganancia_Real'],
                'Ganancia_Mar_EUR':    rent_m['Ganancia_Real'],
                'PVP_Sugerido_EUR':    rent_a['PVP_Sugerido'],
                'Nota':                prov.get('Nota', ''),
                'Link_Alibaba':        prov['Link_Alibaba'],
                'Link_Bol':            prod.get('Link_Bol', ''),
            })

    if not resultados:
        print("⚠️  Sin resultados para exportar.")
        return

    df = pd.DataFrame(resultados)

    # Ordenar: GANADOR primero, luego por ganancia desc
    orden = {'GANADOR': 0, 'POTENCIAL': 1, 'MARGINAL': 2, 'N/D': 3}
    df['_ord'] = df['Estado'].map(orden).fillna(3)
    df = (df.sort_values(['_ord', 'Ganancia_Aereo_EUR'], ascending=[True, False])
            .drop(columns='_ord')
            .reset_index(drop=True))

    # ── 3. Export Excel ────────────────────────────────────────────────────────
    with pd.ExcelWriter(ARCHIVO_SALIDA, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Analisis')
        colorear_excel(writer.sheets['Analisis'], df)
        _agregar_leyenda(writer.book)

    g = (df['Estado'] == 'GANADOR').sum()
    p = (df['Estado'] == 'POTENCIAL').sum()
    m = (df['Estado'] == 'MARGINAL').sum()
    n = (df['Estado'] == 'N/D').sum()

    print(f"\n{'=' * 65}")
    print(f"✅ Excel generado: {ARCHIVO_SALIDA}")
    print(f"   🟢 GANADORES:   {g}  (ganancia > €10 vía aéreo)")
    print(f"   🟡 POTENCIAL:   {p}  (ganancia €5–€10)")
    print(f"   🔴 MARGINAL:    {m}  (ganancia < €5)")
    print(f"   ⬜ SIN PRECIO:  {n}  (verificar link Alibaba manualmente)")
    print(f"\n💡 BTW de importación (~21% CIF) excluido del costo —")
    print(f"   es recuperable para importadores registrados en NL.")
    print(f"{'=' * 65}")


# ── Leyenda personalizada ──────────────────────────────────────────────────────
def _agregar_leyenda(wb):
    if 'Leyenda' in wb.sheetnames:
        del wb['Leyenda']
    ws = wb.create_sheet('Leyenda')

    H = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    G = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    Y = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    R = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    N = PatternFill(fill_type=None)

    filas = [
        (H, 'Campo',            'Descripción'),
        (G, 'GANADOR',          'Ganancia neta > €10 vendiendo al precio actual de Bol.com (vía aéreo)'),
        (Y, 'POTENCIAL',        'Ganancia neta €5–€10. Vale la pena evaluar.'),
        (R, 'MARGINAL',         'Ganancia < €5. El precio de Bol es demasiado bajo o el costo es alto.'),
        (N, '',                 ''),
        (N, 'Precio_FOB_USD',   'Precio del proveedor en Alibaba (Free On Board — sin flete)'),
        (N, 'Precio_FOB_EUR',   'FOB convertido a EUR al tipo de cambio del día'),
        (N, 'Peso_Est_kg',      'Peso estimado por categoría. Ajustar si se conoce el dato real.'),
        (N, 'Flete_x_Unit_Aereo','Air cargo China → NL: €6.50/kg × MOQ ÷ MOQ. Mín €15/envío. Flete por unidad.'),
        (N, 'Flete_x_Unit_Mar', 'LCL sea freight China → Rotterdam: €2.50/kg × MOQ ÷ MOQ. Mín €40/envío.'),
        (N, 'Flete_x_Unit_Expr','DHL/FedEx: €16/kg × MOQ ÷ MOQ. Mín €25/envío. Solo para muestras.'),
        (N, 'Arancel_EUR',      'Arancel EU de importación según categoría (0% electrónica, ~4.7% juguetes…)'),
        (N, 'BTW_Import_EUR',   '⚠️ IVA 21% sobre CIF+arancel. RECUPERABLE para importadores registrados en NL.'),
        (N, 'Costo_Unit_Aereo', 'FOB + Flete Aéreo + Arancel (sin BTW). Base de costo real por unidad.'),
        (N, 'Ganancia_Aereo',   'Precio Bol (neto IVA) × 0.88 − €1 comisión fija − Costo_Unit_Aereo'),
        (N, 'PVP_Sugerido',     'Precio mínimo en Bol para obtener €5 ganancia neta vía flete aéreo'),
        (N, 'Gold_Supplier',    '✓ = proveedor Gold Supplier en Alibaba (verificado ≥1 año)'),
        (N, 'Trade_Assurance',  '✓ = cubierto por Trade Assurance (protección de pagos Alibaba)'),
        (N, 'MOQ',              'Minimum Order Quantity — cantidad mínima por pedido al proveedor'),
    ]

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 72

    for r_idx, (fill, col_a, col_b) in enumerate(filas, 1):
        for c, val in [(1, col_a), (2, col_b)]:
            cell = ws.cell(row=r_idx, column=c, value=val)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True) if fill is H else Font(bold=(r_idx <= 4 and fill is not N))
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[r_idx].height = 22


if __name__ == '__main__':
    ejecutar()
