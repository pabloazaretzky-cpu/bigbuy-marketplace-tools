from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILL_VERDE    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_AMARILLO = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_ROJO     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_HEADER   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_HEADER   = Font(color='FFFFFF', bold=True)
FONT_GANADOR  = Font(bold=True)

FILL_MAP = {
    'GANADOR':   FILL_VERDE,
    'POTENCIAL': FILL_AMARILLO,
    'MARGINAL':  FILL_ROJO,
}

def colorear_excel(ws, df):
    """
    Colors worksheet rows based on the 'Estado' column.
    GANADOR=green, POTENCIAL=yellow, MARGINAL=red.
    Also formats the header row and auto-adjusts column widths.
    """
    # Header styling
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Row coloring
    estado_col_idx = None
    headers = [cell.value for cell in ws[1]]
    if 'Estado' in headers:
        estado_col_idx = headers.index('Estado')

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        estado = row.get('Estado', 'MARGINAL')
        fill = FILL_MAP.get(estado, FILL_ROJO)
        for cell in ws[i]:
            cell.fill = fill
            if estado == 'GANADOR':
                cell.font = FONT_GANADOR

    # Auto column widths (capped at 45)
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    # Freeze header row
    ws.freeze_panes = 'A2'

def agregar_leyenda(wb):
    """Adds a 'Leyenda' sheet explaining the color scheme."""
    if 'Leyenda' in wb.sheetnames:
        del wb['Leyenda']
    ws = wb.create_sheet('Leyenda')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 55

    datos = [
        ('Estado', 'Significado', True),
        ('GANADOR',   'Margen vs Bol > €10 o ganancia neta ≥ €18. Prioridad máxima.', False),
        ('POTENCIAL', 'Margen vs Bol entre €0-€10 o ganancia ≥ €12. Vale la pena evaluar.', False),
        ('MARGINAL',  'Sin margen competitivo o ganancia < €12. Riesgo alto.', False),
    ]
    fills = [FILL_HEADER, FILL_VERDE, FILL_AMARILLO, FILL_ROJO]
    fonts = [FONT_HEADER, FONT_GANADOR, Font(), Font()]

    for row_idx, ((col_a, col_b, _), fill, font) in enumerate(zip(datos, fills, fonts), start=1):
        for col, val in [(1, col_a), (2, col_b)]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True) if row_idx == 1 else font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 25
