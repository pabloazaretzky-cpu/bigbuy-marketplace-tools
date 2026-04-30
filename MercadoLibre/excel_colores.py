from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILL_VERDE    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FILL_AMARILLO = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_ROJO     = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_GRIS     = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
FILL_HEADER   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
FONT_HEADER   = Font(color='FFFFFF', bold=True)
FONT_GANADOR  = Font(bold=True)

FILL_MAP = {
    'GANADOR':   FILL_VERDE,
    'POTENCIAL': FILL_AMARILLO,
    'MARGINAL':  FILL_ROJO,
    'SIN_DATOS': FILL_GRIS,
}

def colorear_excel(ws, df):
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30

    for i, (_, row) in enumerate(df.iterrows(), start=2):
        estado = row.get('Estado', 'MARGINAL')
        fill = FILL_MAP.get(estado, FILL_ROJO)
        for cell in ws[i]:
            cell.fill = fill
            if estado == 'GANADOR':
                cell.font = FONT_GANADOR

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 45)

    ws.freeze_panes = 'A2'

def agregar_leyenda(wb):
    if 'Leyenda' in wb.sheetnames:
        del wb['Leyenda']
    ws = wb.create_sheet('Leyenda')
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60

    datos = [
        ('Estado',    'Significado'),
        ('GANADOR',   'Margen neto > €15 sobre el precio ML Mexico. Alta prioridad.'),
        ('POTENCIAL', 'Margen neto €5–€15. Vale la pena evaluar.'),
        ('MARGINAL',  'Margen < €5 o negativo. Riesgo alto.'),
        ('SIN_DATOS', 'No se encontró el producto en ML Mexico.'),
    ]
    fills  = [FILL_HEADER, FILL_VERDE, FILL_AMARILLO, FILL_ROJO, FILL_GRIS]
    header = True

    for (col_a, col_b), fill in zip(datos, fills):
        row_idx = datos.index((col_a, col_b)) + 1
        for col, val in [(1, col_a), (2, col_b)]:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = Font(color='FFFFFF', bold=True) if row_idx == 1 else Font()
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 25
